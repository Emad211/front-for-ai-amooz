"""openpyxl workbook builders for the advisory reports (risman step 2).

One generic entry point — ``report_workbook(kind, report) -> BytesIO`` — that
dispatches to a per-kind builder over the **already-aggregated** wire dicts
from ``services.reports``. This module touches no tenancy state and no ORM:
it only renders what it is handed (pinned on
``test_import_boundaries._EXEMPT_FILES`` for exactly that reason).

Layout rules from the roadmap:

* one sheet per report («برنامه» for the planner), ``sheet_view.rightToLeft``
  set so Excel opens it reading right-to-left like every other Persian surface;
* a styled header row — bold white font on a neutral dark-green fill — with
  sensible column widths;
* the planner sheet carries the per-day block first (تاریخ / روز /
  برنامه‌ریزی‌شده / انجام‌شده / پوشش٪), then a blank separator row and the
  per-subject block with its totals line.

Numbers stay numeric cells (Excel-friendly); dates are ISO strings; a null
coverage («ثبت نشده») renders as an empty cell, never a fake 0%. No Jalali
math happens here (ق۵): the روز column is the Gregorian weekday's Persian
name, which is display-only.
"""

from __future__ import annotations

import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Neutral dark-green brand-adjacent header fill (ARGB). One constant so the
# export can never drift from itself between sheets or kinds.
HEADER_FILL_HEX = 'FF1F7A5C'

_HEADER_FONT = Font(bold=True, color='FFFFFFFF')
_HEADER_FILL = PatternFill(fill_type='solid', fgColor=HEADER_FILL_HEX)
_HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

# Python's weekday(): Monday=0 .. Sunday=6. The Iranian week starts شنبه.
_WEEKDAY_NAMES_FA = {
    5: 'شنبه',
    6: 'یکشنبه',
    0: 'دوشنبه',
    1: 'سه‌شنبه',
    2: 'چهارشنبه',
    3: 'پنج‌شنبه',
    4: 'جمعه',
}

_PLANNER_DAY_HEADERS = ['تاریخ', 'روز', 'برنامه‌ریزی‌شده', 'انجام‌شده', 'پوشش٪']
_PLANNER_SUBJECT_HEADERS = ['درس', '', 'برنامه‌ریزی‌شده', 'انجام‌شده', 'پوشش٪']

# Column widths (Excel units) — wide enough for the Persian headers not to clip.
_PLANNER_COLUMN_WIDTHS = {1: 14, 2: 12, 3: 18, 4: 14, 5: 10}


def _persian_weekday(day: datetime.date) -> str:
    """The Gregorian date's weekday name in Persian (display-only, ق۵-safe)."""
    return _WEEKDAY_NAMES_FA.get(day.weekday(), '')


def _style_header_row(sheet, row_index: int, column_count: int) -> None:
    """Bold white-on-dark-green centered header across ``column_count`` cells."""
    for column in range(1, column_count + 1):
        cell = sheet.cell(row=row_index, column=column)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT


def _coverage_cell(coverage_percent) -> int | str:
    """Numeric coverage when measured; empty cell when quiet-null."""
    if coverage_percent is None:
        return ''
    return coverage_percent


def planner_workbook(report: dict) -> io.BytesIO:
    """Render a ``planner_report`` payload as the «برنامه» sheet.

    Row order mirrors the wire: per-day rows ascending, then the subject block
    (name-sorted, as the service emits it) closed by a bolded جمع totals row.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'برنامه'
    sheet.sheet_view.rightToLeft = True

    for column, width in _PLANNER_COLUMN_WIDTHS.items():
        sheet.column_dimensions[chr(64 + column)].width = width

    # ── per-day block ────────────────────────────────────────────────────────
    sheet.append(_PLANNER_DAY_HEADERS)
    _style_header_row(sheet, 1, len(_PLANNER_DAY_HEADERS))
    for day in report.get('days', []):
        sheet.append([
            day['date'],
            _persian_weekday(datetime.date.fromisoformat(day['date'])),
            day['planned'],
            day['actual'],
            _coverage_cell(day.get('coveragePercent')),
        ])

    # Blank separator row between the two blocks.
    sheet.append([])

    # ── per-subject block ────────────────────────────────────────────────────
    subject_header_row = sheet.max_row + 1
    sheet.append(_PLANNER_SUBJECT_HEADERS)
    _style_header_row(sheet, subject_header_row, len(_PLANNER_SUBJECT_HEADERS))
    for subject in report.get('subjects', []):
        sheet.append([
            subject['name'],
            '',
            subject['planned'],
            subject['actual'],
            _coverage_cell(subject.get('coveragePercent')),
        ])

    totals = report.get('totals') or {}
    sheet.append([
        'جمع',
        '',
        totals.get('planned', 0),
        totals.get('actual', 0),
        _coverage_cell(totals.get('coveragePercent')),
    ])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


_ORG_ADVISOR_HEADERS = [
    'مشاور', 'دانش‌آموزان', 'برنامه‌ریزی‌شده', 'انجام‌شده', 'پوشش٪',
    'برنامه‌های منتشرشده', 'ارزیابی‌های هفتگی', 'تحلیل‌های آزمون',
]
_ORG_STUDENT_HEADERS = [
    'مشاور', 'دانش‌آموز', 'برنامه‌ریزی‌شده', 'انجام‌شده', 'پوشش٪', 'آزمون',
]

# Column widths (Excel units) — the summary block owns columns 1..8, the
# student block the first six of them.
_ORG_COLUMN_WIDTHS = {1: 20, 2: 18, 3: 16, 4: 14, 5: 10, 6: 18, 7: 16, 8: 14}


def org_advisor_workbook(report: dict) -> io.BytesIO:
    """Render the org-manager per-advisor report (risman step 3) as one sheet.

    Mirrors the planner layout: the per-advisor summary block first, then a
    blank separator and the flattened per-student block (one row per student
    with their advisor's name beside it). Numbers stay numeric; a quiet-null
    coverage («ثبت نشده») renders as an empty cell, never a fake 0%.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'مشاوران'
    sheet.sheet_view.rightToLeft = True

    for column, width in _ORG_COLUMN_WIDTHS.items():
        sheet.column_dimensions[chr(64 + column)].width = width

    # ── per-advisor summary block ────────────────────────────────────────────
    sheet.append(_ORG_ADVISOR_HEADERS)
    _style_header_row(sheet, 1, len(_ORG_ADVISOR_HEADERS))
    for advisor in report.get('advisors', []):
        sheet.append([
            advisor['advisorName'],
            advisor['studentCount'],
            advisor['planned'],
            advisor['actual'],
            _coverage_cell(advisor.get('coveragePercent')),
            advisor['plansPublished'],
            advisor['assessmentsWritten'],
            advisor['analysesCreated'],
        ])

    # Blank separator row between the two blocks.
    sheet.append([])

    # ── per-student block ────────────────────────────────────────────────────
    student_header_row = sheet.max_row + 1
    sheet.append(_ORG_STUDENT_HEADERS)
    _style_header_row(sheet, student_header_row, len(_ORG_STUDENT_HEADERS))
    for advisor in report.get('advisors', []):
        for student in advisor.get('students', []):
            sheet.append([
                advisor['advisorName'],
                student['studentName'],
                student['planned'],
                student['actual'],
                _coverage_cell(student.get('coveragePercent')),
                student['testsTaken'],
            ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


_BUILDERS = {
    'planner': planner_workbook,
}


def report_workbook(kind: str, report: dict) -> io.BytesIO:
    """Build the xlsx bytes for ``kind`` from an aggregated report dict."""
    builder = _BUILDERS.get(kind)
    if builder is None:
        raise ValueError(f'Unknown report kind: {kind!r}')
    return builder(report)
