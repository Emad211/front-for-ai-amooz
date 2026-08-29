"""Risman step 3 — the thin org-manager panel: overview, advisor report, reassign.

Zero-token, ORM-fixture arithmetic only (roadmap ق۴). The doctrine applied:

* **exact math** — every counter in the overview and every cell of the advisor
  report is hand-derived from fixtures (plans at offset 0 / elapsed offsets,
  logs at TODAY, one draft excluded, one foreign-engagement excluded), so a
  regression in the shared bucket helpers shows up as a wrong number, not a
  crash;
* **tenancy** — a manager with no (ACTIVE admin/deputy) membership is a
  stranger and gets 404 «سازمانی برای شما یافت نشد.» on all three routes; a
  freelance engagement of the same humans and a second org's data are both
  invisible to the numbers;
* **the single write** — reassign rules are exercised one violation at a time,
  each with its pinned Persian message, plus the happy path's audit row;
* **full access matrix** — advisor/student/teacher 403, anonymous 401,
  manager-of-nowhere 404, deputy 200;
* **Excel branch** — the report under ``?format=xlsx`` is read back with
  openpyxl: sheet title, RTL flag, header cells and the advisor rows.
"""

from __future__ import annotations

import datetime
import io

import pytest
from django.contrib.auth import get_user_model
from django.utils import timezone
from model_bakery import baker
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.state import token_backend
from rest_framework_simplejwt.tokens import RefreshToken

from apps.advisory.models import (
    AdvisoryAccessLog,
    AdvisoryEngagement,
    DailyLog,
    DailyLogItem,
    StudyExamAnalysis,
    StudyPlan,
    StudyPlanItem,
    StudentSubject,
    Subject,
    WeeklyAssessment,
)
from apps.advisory.services import calendar
from apps.organizations.models import (
    ImpersonationLog,
    Organization,
    OrganizationMembership,
)

User = get_user_model()
Status = AdvisoryEngagement.Status
Mode = AdvisoryEngagement.Mode
OrgRole = OrganizationMembership.OrgRole
MStatus = OrganizationMembership.MemberStatus

pytestmark = [pytest.mark.django_db, pytest.mark.api]

ORG_OVERVIEW = '/api/advisory/org/overview/'
ORG_REPORT = '/api/advisory/org/advisors/'
REASSIGN = '/api/advisory/org/engagements/{pk}/reassign/'

MSG_NO_ORG = 'سازمانی برای شما یافت نشد.'
MSG_BAD_DATE = 'تاریخ باید به شکل YYYY-MM-DD باشد.'
MSG_REVERSED_RANGE = 'تاریخ پایان نمی‌تواند پیش از تاریخ شروع باشد.'
MSG_RANGE_TOO_LONG = 'بازه حداکثر ۹۲ روز است.'
MSG_BODY = 'بدنه‌ی درخواست نامعتبر است.'

TODAY = timezone.localdate()
WEEK_START = calendar.week_start_of(TODAY)
FROM = TODAY - datetime.timedelta(days=6)
TO = TODAY


# ── fixture helpers ───────────────────────────────────────────────────────────

def _auth(user) -> APIClient:
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def _org(name='دبیرستان البرز', **kwargs) -> Organization:
    kwargs.setdefault('subscription_status', Organization.SubscriptionStatus.ACTIVE)
    return baker.make(Organization, name=name, **kwargs)


def _member(org, org_role, username=None, **user_kwargs):
    """A user + ACTIVE membership row of the given org role."""
    platform_role = {
        OrgRole.ADMIN: User.Role.MANAGER,
        OrgRole.DEPUTY: User.Role.MANAGER,
        OrgRole.ADVISOR: User.Role.ADVISOR,
        OrgRole.STUDENT: User.Role.STUDENT,
    }[org_role]
    user = baker.make(
        User, username=username or f'u-{org_role}-{User.objects.count()}',
        role=platform_role, **user_kwargs,
    )
    baker.make(
        OrganizationMembership, user=user, organization=org,
        org_role=org_role, status=MStatus.ACTIVE,
    )
    return user


def _student(username='stu', phone=None, **kwargs):
    """A STUDENT shell. ``phone`` defaults to None (bakery leaves it NULL) so
    two shells in one transaction never collide on the unique-phone index."""
    if phone is not None:
        kwargs['phone'] = phone
    return baker.make(User, username=username, role=User.Role.STUDENT, **kwargs)


def _org_engagement(advisor, student, org, *, status_value=Status.ACTIVE, **kwargs):
    defaults = {
        'invited_phone': student.phone or '',
        'mode': Mode.ORG,
        'organization': org,
        'status': status_value,
        'started_on': TODAY - datetime.timedelta(days=60),
    }
    defaults.update(kwargs)
    return AdvisoryEngagement.objects.create(
        advisor=advisor, student=student, **defaults,
    )


def _subject(name: str) -> Subject:
    return baker.make(Subject, name=name)


def _selection(engagement, subject) -> StudentSubject:
    return baker.make(StudentSubject, engagement=engagement, subject=subject)


def _make_plan(engagement, start, duration_days, rows, status_value) -> StudyPlan:
    """rows = [(day_offset, selection, planned_minutes)] — ORM-direct, no door."""
    plan = baker.make(
        StudyPlan, engagement=engagement, start_date=start,
        duration_days=duration_days, status=status_value,
    )
    StudyPlanItem.objects.bulk_create([
        StudyPlanItem(
            plan=plan, day_offset=offset, student_subject=sel,
            planned_minutes=minutes,
        )
        for offset, sel, minutes in rows
    ])
    return plan


def _log(engagement, log_date, rows, tests_taken=0) -> DailyLog:
    """rows = [(selection, actual_minutes)] — ORM-direct, no door."""
    log = baker.make(
        DailyLog, engagement=engagement, log_date=log_date,
        tests_taken=tests_taken or None,
    )
    DailyLogItem.objects.bulk_create([
        DailyLogItem(log=log, student_subject=sel, actual_minutes=minutes)
        for sel, minutes in rows
    ])
    return log


# ── the access matrix ─────────────────────────────────────────────────────────

@pytest.fixture
def org():
    """A bare ACTIVE organization for the access-matrix cells."""
    return _org()


class TestAccessMatrix:
    """Role + tenancy gates of all three routes, one assertion per cell."""

    def test_anonymous_gets_401_on_overview(self):
        assert APIClient().get(ORG_OVERVIEW).status_code == 401

    def test_advisor_gets_403_on_overview(self, org):
        res = _auth(_member(org, OrgRole.ADVISOR)).get(ORG_OVERVIEW)
        assert res.status_code == 403
        assert res.json()['detail'] == 'فقط مدیر موسسه اجازه دسترسی دارد.'

    def test_student_gets_403_on_report(self, org):
        res = _auth(_member(org, OrgRole.STUDENT)).get(
            f'{ORG_REPORT}?from={FROM}&to={TO}',
        )
        assert res.status_code == 403
        assert res.json()['detail'] == 'فقط مدیر موسسه اجازه دسترسی دارد.'

    def test_teacher_role_gets_403_on_reassign(self, org):
        teacher = baker.make(User, username='tch', role=User.Role.TEACHER)
        res = _auth(teacher).post(REASSIGN.format(pk=1), {'advisorId': 1}, format='json')
        assert res.status_code == 403

    def test_manager_without_membership_is_a_stranger_404(self):
        """Valid MANAGER, but oversees nothing ⇒ 404 on every route، نه ۴۰۳."""
        manager = baker.make(User, username='lone', role=User.Role.MANAGER)
        client = _auth(manager)
        assert client.get(ORG_OVERVIEW).status_code == 404
        assert client.get(f'{ORG_REPORT}?from={FROM}&to={TO}').status_code == 404
        res = client.post(REASSIGN.format(pk=1), {'advisorId': 1}, format='json')
        assert res.status_code == 404
        assert res.json()['detail'] == MSG_NO_ORG

    def test_suspended_membership_loses_the_panel(self, org):
        manager = _member(org, OrgRole.ADMIN)
        OrganizationMembership.objects.filter(user=manager).update(status=MStatus.SUSPENDED)
        assert _auth(manager).get(ORG_OVERVIEW).status_code == 404

    def test_expired_org_subscription_goes_dark(self, org):
        manager = _member(org, OrgRole.ADMIN)
        Organization.objects.filter(pk=org.pk).update(
            subscription_status=Organization.SubscriptionStatus.EXPIRED,
        )
        assert _auth(manager).get(ORG_OVERVIEW).status_code == 404

    def test_deputy_is_accepted(self, org):
        deputy = _member(org, OrgRole.DEPUTY)
        assert _auth(deputy).get(ORG_OVERVIEW).status_code == 200

    def test_platform_admin_is_strictly_excluded(self, org):
        admin = baker.make(User, username='root', role=User.Role.ADMIN, is_staff=True)
        assert _auth(admin).get(ORG_OVERVIEW).status_code == 403


# ── shared fixture: one org with two advisors and three students ─────────────

@pytest.fixture
def panel():
    """Org A staffed for math checks, plus freelance and foreign-org noise.

    Org A (البرز): advisor «آرش کمالی» with students «زهرا مرادی» and «سارا
    احمدی», advisor «مریم رضایی» with student «علی کریمی». All engagements ORG
    + ACTIVE, started 60 days ago. Returns a dict the tests count over.
    """
    org = _org()
    manager = _member(org, OrgRole.ADMIN, 'boss', first_name='مدیر', last_name='برز')
    adv1 = _member(org, OrgRole.ADVISOR, 'adv1', first_name='آرش', last_name='کمالی')
    adv2 = _member(org, OrgRole.ADVISOR, 'adv2', first_name='مریم', last_name='رضایی')
    s1 = _member(org, OrgRole.STUDENT, 'stu1', first_name='زهرا', last_name='مرادی', phone='09120000001')
    s2 = _member(org, OrgRole.STUDENT, 'stu2', first_name='سارا', last_name='احمدی', phone='09120000002')
    s3 = _member(org, OrgRole.STUDENT, 'stu3', first_name='علی', last_name='کریمی', phone='09120000003')

    eng1 = _org_engagement(adv1, s1, org)   # the fully instrumented pair
    eng2 = _org_engagement(adv1, s2, org)   # silent student
    eng3 = _org_engagement(adv2, s3, org)   # planned but not studying

    return {
        'org': org, 'manager': manager, 'adv1': adv1, 'adv2': adv2,
        's1': s1, 's2': s2, 's3': s3,
        'eng1': eng1, 'eng2': eng2, 'eng3': eng3,
    }


# ── org_overview: exact counter arithmetic ────────────────────────────────────

class TestOverviewMath:
    """Every headline number hand-derived; noise must not move them."""

    def test_exact_counters(self, panel):
        p = panel
        math_s = _selection(p['eng1'], _subject('ریاضی'))
        phys_s = _selection(p['eng1'], _subject('فیزیک'))

        # Planned this week, both rows at offset 0 so the answer cannot depend
        # on how much of the week has elapsed when the suite runs.
        _make_plan(
            p['eng1'], WEEK_START, 7,
            [(0, math_s, 120), (0, phys_s, 30)], StudyPlan.Status.PUBLISHED,
        )
        # A DRAFT plan of the same week never counts as «منتشرشده».
        _make_plan(p['eng2'], WEEK_START, 7, [], StudyPlan.Status.DRAFT)
        eng3_math = _selection(p['eng3'], _subject('شیمی'))
        _make_plan(p['eng3'], WEEK_START, 7, [(0, eng3_math, 45)], StudyPlan.Status.PUBLISHED)

        # Studied today: 80 real minutes against 150 planned for eng1.
        _log(p['eng1'], TODAY, [(math_s, 60), (phys_s, 20)], tests_taken=5)

        res = _auth(p['manager']).get(ORG_OVERVIEW)
        assert res.status_code == 200
        assert res.json() == {
            'activeStudents': 3,
            'activeAdvisors': 2,
            'activeEngagements': 3,
            'weekPlansPublished': 2,
            'logsToday': 1,
            'minutesToday': 80,
            # Weighted overall: round(80 / (150+45) × 100) = 41 — not an
            # average of per-student percentages.
            'avgCommitmentPercent': 41,
        }

    def test_freelance_engagement_of_same_advisor_is_invisible(self, panel):
        """ق۳ literally: advisory-org endpoints must ignore freelance work."""
        p = panel
        freelance_stu = _student('free1')
        free_eng = AdvisoryEngagement.objects.create(
            advisor=p['adv1'], student=freelance_stu,
            invited_phone=freelance_stu.phone or '',
            mode=Mode.FREELANCE, organization=None, status=Status.ACTIVE,
            started_on=TODAY - datetime.timedelta(days=60),
        )
        sel = _selection(free_eng, _subject('ادبیات'))
        _make_plan(free_eng, TODAY - datetime.timedelta(days=30), 7,
                   [(0, sel, 300)], StudyPlan.Status.PUBLISHED)
        _log(free_eng, TODAY, [(sel, 300)], tests_taken=99)

        body = _auth(p['manager']).get(ORG_OVERVIEW).json()
        assert body['activeEngagements'] == 3
        # The freelance plan+log of TODAY exist, but the org counters must
        # stay clean zeroes — the noise never leaks in.
        assert body['logsToday'] == 0
        assert body['minutesToday'] == 0

    def test_other_org_data_is_invisible(self, panel):
        """The classic IDOR in aggregate form — org B's numbers stay org B's."""
        p = panel
        other = _org(name='دبیرستان رازی')
        o_adv = _member(other, OrgRole.ADVISOR)
        o_stu = _member(other, OrgRole.STUDENT, phone='09120000999')
        o_eng = _org_engagement(o_adv, o_stu, other)
        sel = _selection(o_eng, _subject('هندسه'))
        _make_plan(o_eng, WEEK_START, 7, [(0, sel, 500)], StudyPlan.Status.PUBLISHED)
        _log(o_eng, TODAY, [(sel, 400)], tests_taken=10)

        body = _auth(p['manager']).get(ORG_OVERVIEW).json()
        assert body['activeStudents'] == 3      # نه ۴
        assert body['activeAdvisors'] == 2      # نه ۳
        assert body['activeEngagements'] == 3   # نه ۴
        assert body['weekPlansPublished'] == 0
        assert body['logsToday'] == 0
        assert body['minutesToday'] == 0

    def test_zero_planned_week_answers_null_not_fake_zero(self, panel):
        """Nothing planned anywhere ⇒ coverage is None، نه ۰٪."""
        res = _auth(panel['manager']).get(ORG_OVERVIEW).json()
        assert res['avgCommitmentPercent'] is None


# ── org_advisor_report: per-advisor table, ranges and the Excel branch ───────

def _instrument_panel(panel):
    """The report-scenario fixtures: plans/logs/tools inside the 7-day window."""
    p = panel
    math_s = _selection(p['eng1'], _subject('ریاضی'))
    eng3_math = _selection(p['eng3'], _subject('شیمی'))

    # s1: PUBLISHED plan whose only counted row lands on TODAY (offset 6 of a
    # plan starting FROM) → planned 100؛ plus a DRAFT 500 that must not count.
    _make_plan(p['eng1'], FROM, 7, [(6, math_s, 100)], StudyPlan.Status.PUBLISHED)
    _make_plan(p['eng1'], FROM, 7, [], StudyPlan.Status.DRAFT)
    _log(p['eng1'], TODAY, [(math_s, 70)], tests_taken=5)

    # s3: PUBLISHED plan whose counted row lands on TODAY (offset 6 of a plan
    # starting FROM) → planned 200 with zero actual ⇒ coverage 0، نه null؛
    # و چون start_date داخل بازه است، plansPublished هم یک می‌شود.
    _make_plan(p['eng3'], FROM, 7,
               [(6, eng3_math, 200)], StudyPlan.Status.PUBLISHED)
    baker.make(
        WeeklyAssessment, engagement=p['eng3'],
        week_start=WEEK_START, scores={'focus': 4},
    )
    baker.make(StudyExamAnalysis, engagement=p['eng3'])
    return p


class TestOrgAdvisorReport:
    """GET org/advisors/ — math, ordering, windows, Excel."""

    def test_exact_rows_and_ordering(self, panel):
        _instrument_panel(panel)
        res = _auth(panel['manager']).get(f'{ORG_REPORT}?from={FROM}&to={TO}')
        assert res.status_code == 200
        advisors = res.json()['advisors']
        assert len(advisors) == 2

        first, second = advisors
        assert first['advisorName'] == 'آرش کمالی'      # 2 students sorts first
        assert second['advisorName'] == 'مریم رضایی'

        assert first['studentCount'] == 2
        assert first['planned'] == 100                  # draft's 500 excluded
        assert first['actual'] == 70
        assert first['coveragePercent'] == 70           # round(70/100×100)
        assert first['plansPublished'] == 1
        assert first['assessmentsWritten'] == 0
        assert first['analysesCreated'] == 0

        [row_s1, row_s2] = sorted(first['students'], key=lambda r: r['studentName'])
        assert row_s2 == {
            'engagementId': panel['eng2'].pk,
            'studentName': 'سارا احمدی',
            'planned': 0, 'actual': 0,
            'coveragePercent': None, 'testsTaken': 0,
        }
        assert row_s1 == {
            'engagementId': panel['eng1'].pk,
            'studentName': 'زهرا مرادی',
            'planned': 100, 'actual': 70,
            'coveragePercent': 70, 'testsTaken': 5,
        }

        assert second['studentCount'] == 1
        assert second['planned'] == 200
        assert second['actual'] == 0
        assert second['coveragePercent'] == 0
        assert second['plansPublished'] == 1
        assert second['assessmentsWritten'] == 1        # this week's Saturday row
        assert second['analysesCreated'] == 1

    def test_reversed_range_message(self, panel):
        res = _auth(panel['manager']).get(
            f'{ORG_REPORT}?from={TO}&to={FROM}',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == MSG_REVERSED_RANGE

    def test_overlong_range_message(self, panel):
        over_from = datetime.date(2026, 1, 1)
        over_to = datetime.date(2026, 4, 3)   # inclusive span = 93 days
        res = _auth(panel['manager']).get(
            f'{ORG_REPORT}?from={over_from}&to={over_to}',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == MSG_RANGE_TOO_LONG

    def test_missing_params_message(self, panel):
        res = _auth(panel['manager']).get(ORG_REPORT)
        assert res.status_code == 400
        assert res.json()['detail'] == MSG_BAD_DATE

    def test_malformed_date_message(self, panel):
        res = _auth(panel['manager']).get(f'{ORG_REPORT}?from=۱۴۰۵&to={TO}')
        assert res.status_code == 400
        assert res.json()['detail'] == MSG_BAD_DATE

    def test_excel_branch_is_read_back(self, panel):
        _instrument_panel(panel)
        res = _auth(panel['manager']).get(f'{ORG_REPORT}?from={FROM}&to={TO}&format=xlsx')
        assert res.status_code == 200
        assert res['Content-Type'].startswith(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        assert res['Content-Disposition'] == (
            f'attachment; filename=report-advisors-{FROM}_{TO}.xlsx'
        )

        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        assert ws.title == 'مشاوران'
        assert ws.sheet_view.rightToLeft is True
        # Advisor summary block: header at row 1, one row per advisor below.
        assert ws['A1'].value == 'مشاور'
        assert ws['B1'].value == 'دانش‌آموزان'
        column_a = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert 'آرش کمالی' in column_a
        assert 'مریم رضایی' in column_a
        # Student block: one row per student, their advisor named beside them.
        column_b = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert 'زهرا مرادی' in column_b
        assert 'علی کریمی' in column_b


# ── permission-class units for the upcoming impersonation guard ───────────────

class TestImpersonationClaimGuard:
    """IsOrgManager rejects tokens bearing the ``imp`` claim (step-4 foreplay).

    A full JWT round-trip belongs to step 4's own suite; a minimal token
    stand-in proves both sides of the claim check here.
    """

    @staticmethod
    def _request_with_payload(user, payload):
        from rest_framework.request import Request
        from rest_framework.test import APIRequestFactory

        raw = APIRequestFactory().get('/api/advisory/org/overview/')
        request = Request(raw)
        request.user = user
        if payload is not None:
            request.auth = type('_Token', (), {'payload': payload})()
        else:
            request.auth = None
        return request

    def test_impersonated_token_is_rejected(self, panel):
        from apps.core.permissions import IsOrgManager

        request = self._request_with_payload(
            panel['manager'], {'imp': {'by': 7, 'org': 1}},
        )
        assert IsOrgManager().has_permission(request, view=None) is False

    def test_clean_manager_token_is_accepted(self, panel):
        from apps.core.permissions import IsOrgManager

        request = self._request_with_payload(
            panel['manager'], {'token_type': 'access'},
        )
        assert IsOrgManager().has_permission(request, view=None) is True


# ── impersonation: minting and spending the 30-minute twin token ─────────────

IMPERSONATE = '/api/organizations/{org}/impersonate/{user}/'
IMPERSONATE_STOP = '/api/organizations/{org}/impersonate/stop/'

MSG_STAFF_ONLY = 'این جلسه در حالت ورود مستقیم است و این کار مجاز نیست.'
MSG_SELF = 'خودتان را نمی‌توانید انتخاب کنید.'
MSG_MEMBER_ONLY = 'هدف باید مشاور یا دانش‌آموز همین سازمان باشد.'
MSG_NOT_MEMBER = 'کاربر موردنظر در این سازمان یافت نشد.'
MSG_NO_OPEN_SESSION = 'جلسه‌ی ورود مستقیمی باز نیست.'


def _bearer(user, extra_payload=None) -> APIClient:
    """A client whose Bearer carries a REAL SimpleJWT access token.

    ``force_authenticate`` cannot represent an impersonation session (there is
    no token object to bear the ``imp`` claim), so every guard against the
    claim itself must be tested with genuinely signed tokens.
    """
    refresh = RefreshToken.for_user(user)
    for key, value in (extra_payload or {}).items():
        refresh.payload[key] = value
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {refresh.access_token}')
    return api


def _refresh_only(user, extra_payload=None) -> APIClient:
    """A client sending the REFRESH token itself as its bearer credential."""
    refresh = RefreshToken.for_user(user)
    for key, value in (extra_payload or {}).items():
        refresh.payload[key] = value
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {refresh}')
    return api


class TestImpersonate:
    """Minting the twin token — security-sensitive paths only."""

    def test_mint_success_carries_imp_claim_and_logs_row(self, panel):
        p = panel
        res = _auth(p['manager']).post(IMPERSONATE.format(org=p['org'].pk, user=p['adv1'].pk))
        assert res.status_code == 200
        body = res.json()
        # Exactly three keys: the client swaps its stored pair wholesale.
        assert set(body.keys()) == {'access', 'refresh', 'user'}
        assert body['user']['id'] == p['adv1'].pk
        assert body['user']['role'] == 'ADVISOR'
        expected_imp = {'by': p['manager'].pk, 'org': p['org'].pk}
        access_claims = token_backend.decode(body['access'], verify=True)
        refresh_claims = token_backend.decode(body['refresh'], verify=True)
        assert access_claims['imp'] == expected_imp
        assert refresh_claims['imp'] == expected_imp

        log = ImpersonationLog.objects.get(
            manager=p['manager'], organization=p['org'], target_user=p['adv1'],
        )
        assert log.started_at is not None
        assert log.ended_at is None

    def test_anonymous_gets_401(self, panel):
        res = APIClient().post(
            IMPERSONATE.format(org=panel['org'].pk, user=panel['adv1'].pk),
        )
        assert res.status_code == 401

    def test_advisor_cannot_mint_for_anyone(self, panel):
        p = panel
        res = _auth(p['adv2']).post(IMPERSONATE.format(org=p['org'].pk, user=p['adv1'].pk))
        assert res.status_code == 403
        assert res.json()['detail'] == 'فقط مدیر موسسه اجازه دسترسی دارد.'

    def test_manager_of_another_org_gets_404_not_403(self, panel):
        """Manager B naming org A's member id — closed failure, no oracle."""
        p = panel
        stranger_manager = _member(_org(name='دبیرستان رازی'), OrgRole.ADMIN)
        res = _auth(stranger_manager).post(
            IMPERSONATE.format(org=p['org'].pk, user=p['s1'].pk),
        )
        assert res.status_code == 404
        assert res.json()['detail'] == MSG_NO_ORG

    def test_self_impersonation_is_rejected(self, panel):
        p = panel
        res = _auth(p['manager']).post(IMPERSONATE.format(org=p['org'].pk, user=p['manager'].pk))
        assert res.status_code == 400
        assert res.json()['detail'] == MSG_SELF

    def test_teacher_with_membership_is_rejected_pinned(self, panel):
        p = panel
        teacher = baker.make(User, username='tch9', role=User.Role.TEACHER)
        baker.make(
            OrganizationMembership, user=teacher, organization=p['org'],
            org_role=OrgRole.TEACHER, status=MStatus.ACTIVE,
        )
        res = _auth(p['manager']).post(IMPERSONATE.format(org=p['org'].pk, user=teacher.pk))
        assert res.status_code == 400
        assert res.json()['detail'] == MSG_MEMBER_ONLY

    def test_user_without_membership_in_this_org_is_404(self, panel):
        p = panel
        outsider = _student('out1')
        res = _auth(p['manager']).post(IMPERSONATE.format(org=p['org'].pk, user=outsider.pk))
        assert res.status_code == 404
        assert res.json()['detail'] == MSG_NOT_MEMBER

    def test_suspended_target_membership_is_not_a_valid_target(self, panel):
        p = panel
        OrganizationMembership.objects.filter(user=p['s1']).update(status=MStatus.SUSPENDED)
        res = _auth(p['manager']).post(IMPERSONATE.format(org=p['org'].pk, user=p['s1'].pk))
        assert res.status_code == 404
        assert res.json()['detail'] == MSG_NOT_MEMBER

    def test_suspended_manager_cannot_mint(self, panel):
        p = panel
        OrganizationMembership.objects.filter(user=p['manager']).update(status=MStatus.SUSPENDED)
        res = _auth(p['manager']).post(IMPERSONATE.format(org=p['org'].pk, user=p['adv1'].pk))
        assert res.status_code == 404
        assert res.json()['detail'] == MSG_NO_ORG


class TestImpersonatedGuards:
    """Spending the twin token — what it must NOT be able to do."""

    def _mint(self, manager, org, target) -> dict:
        return _auth(manager).post(
            IMPERSONATE.format(org=org.pk, user=target.pk),
        ).json()

    def test_imp_bearer_blocked_at_permission_layer(self, panel):
        """The token's OWN view: the derived access token hits org overview."""
        p = panel
        tokens = self._mint(p['manager'], p['org'], p['adv1'])
        api = APIClient()
        api.credentials(HTTP_AUTHORIZATION=f"Bearer {tokens['access']}")
        response = api.get(ORG_OVERVIEW)
        assert response.status_code == 403
        assert response.json()['detail'] == MSG_STAFF_ONLY

    def test_imp_session_cannot_mint_another_impersonation(self, panel):
        p = panel
        tokens = self._mint(p['manager'], p['org'], p['adv1'])
        api = APIClient()
        api.credentials(HTTP_AUTHORIZATION=f"Bearer {tokens['access']}")
        res = api.post(IMPERSONATE.format(org=p['org'].pk, user=p['s1'].pk))
        assert res.status_code == 403
        assert res.json()['detail'] == MSG_STAFF_ONLY

    def test_stop_registers_end_once_then_the_session_is_gone(self, panel):
        p = panel
        tokens = self._mint(p['manager'], p['org'], p['adv1'])
        stop = _refresh_only(p['manager']).post(
            IMPERSONATE_STOP.format(org=p['org'].pk),
        )
        assert stop.status_code == 200
        assert stop.json() == {'ended': True}
        log = ImpersonationLog.objects.get(manager=p['manager'], organization=p['org'])
        assert log.target_user_id == p['adv1'].pk
        assert log.ended_at is not None

        again = _refresh_only(p['manager']).post(
            IMPERSONATE_STOP.format(org=p['org'].pk),
        )
        assert again.status_code == 404
        assert again.json()['detail'] == MSG_NO_OPEN_SESSION

    def test_stop_requires_refresh_class_token(self, panel):
        """The SAME manager with the derived ACCESS token cannot end sessions."""
        p = panel
        self._mint(p['manager'], p['org'], p['adv1'])
        res = _bearer(p['manager']).post(IMPERSONATE_STOP.format(org=p['org'].pk))
        assert res.status_code == 403
        assert res.json()['detail'] == MSG_STAFF_ONLY

    def test_stop_without_open_session_is_404(self, panel):
        p = panel
        res = _refresh_only(p['manager']).post(
            IMPERSONATE_STOP.format(org=p['org'].pk),
        )
        assert res.status_code == 404
        assert res.json()['detail'] == MSG_NO_OPEN_SESSION


# ── reassign: the panel's one write, one violation per test ──────────────────

class TestReassign:
    """POST org/engagements/<pk>/reassign/ — rules and their pinned messages."""

    def test_happy_path_swaps_advisor_and_writes_audit_row(self, panel):
        p = panel
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=p['eng1'].pk),
            {'advisorId': p['adv2'].pk}, format='json',
        )
        assert res.status_code == 200
        body = res.json()
        assert body['engagementId'] == p['eng1'].pk
        assert body['advisorId'] == p['adv2'].pk
        assert body['advisorName'] == 'مریم رضایی'
        assert body['studentName'] == 'زهرا مرادی'

        p['eng1'].refresh_from_db(fields=['advisor'])
        assert p['eng1'].advisor_id == p['adv2'].pk
        log = AdvisoryAccessLog.objects.filter(
            engagement=p['eng1'], action='org_reassign',
        ).get()   # exactly one row — .get() fails loudly on duplicates
        assert log.reader_id == p['manager'].pk

    def test_body_shape_guards(self, panel):
        """Missing key, wrong key, wrong type — all one 400, no partial move."""
        p = panel
        client = _auth(p['manager'])
        res = client.post(REASSIGN.format(pk=p['eng1'].pk), {}, format='json')
        assert res.status_code == 400
        assert res.json()['detail'] == MSG_BODY
        assert client.post(
            REASSIGN.format(pk=p['eng1'].pk), {'x': 1}, format='json',
        ).status_code == 400
        assert client.post(
            REASSIGN.format(pk=p['eng1'].pk), {'advisorId': 'abc'}, format='json',
        ).status_code == 400

    def test_missing_engagement_404(self, panel):
        res = _auth(panel['manager']).post(
            REASSIGN.format(pk=987654),
            {'advisorId': panel['adv1'].pk}, format='json',
        )
        assert res.status_code == 404

    def test_foreign_org_engagement_is_404_not_403(self, panel):
        """ق۶: another org's engagement answers the same blank 404."""
        p = panel
        other = _org(name='دبیرستان رازی')
        o_adv = _member(other, OrgRole.ADVISOR)
        o_stu = _member(other, OrgRole.STUDENT, phone='09120000777')
        other_eng = _org_engagement(o_adv, o_stu, other)

        res = _auth(p['manager']).post(
            REASSIGN.format(pk=other_eng.pk),
            {'advisorId': p['adv1'].pk}, format='json',
        )
        assert res.status_code == 404
        other_eng.refresh_from_db()
        assert other_eng.advisor_id == o_adv.pk      # nothing moved

    def test_freelance_engagement_is_out_of_reach(self, panel):
        """The org gate never touches a freelance pair of the same humans."""
        p = panel
        free_stu = _student('free9')
        free_eng = AdvisoryEngagement.objects.create(
            advisor=p['adv1'], student=free_stu,
            invited_phone=free_stu.phone or '',
            mode=Mode.FREELANCE, organization=None, status=Status.ACTIVE,
            started_on=TODAY - datetime.timedelta(days=30),
        )
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=free_eng.pk),
            {'advisorId': p['adv2'].pk}, format='json',
        )
        assert res.status_code == 404
        free_eng.refresh_from_db()
        assert free_eng.advisor_id == p['adv1'].pk

    def test_non_active_status_refused(self, panel):
        p = panel
        ended = _org_engagement(
            p['adv1'], _student('stu9'), p['org'], status_value=Status.ENDED,
        )
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=ended.pk),
            {'advisorId': p['adv2'].pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'فقط همکاری فعال قابل جابجایی است.'

    def test_teacher_platform_role_refused(self, panel):
        p = panel
        teacher = baker.make(User, username='tch2', role=User.Role.TEACHER)
        baker.make(
            OrganizationMembership, user=teacher, organization=p['org'],
            org_role=OrgRole.TEACHER, status=MStatus.ACTIVE,
        )
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=p['eng1'].pk),
            {'advisorId': teacher.pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'کاربر انتخابی مشاور نیست.'

    def test_suspended_advisor_membership_refused(self, panel):
        p = panel
        extra_adv = _member(p['org'], OrgRole.ADVISOR, 'adv3')
        OrganizationMembership.objects.filter(user=extra_adv).update(status=MStatus.SUSPENDED)
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=p['eng1'].pk),
            {'advisorId': extra_adv.pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'مشاور انتخابی به این سازمان تعلق ندارد.'
        p['eng1'].refresh_from_db()
        assert p['eng1'].advisor_id == p['adv1'].pk

    def test_foreign_org_advisor_membership_refused(self, panel):
        """An advisor whose membership lives in ANOTHER org is not eligible."""
        p = panel
        other = _org(name='دبیرستان رازی')
        foreign_adv = _member(other, OrgRole.ADVISOR, 'foreign')
        before = p['eng1'].advisor_id
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=p['eng1'].pk),
            {'advisorId': foreign_adv.pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'مشاور انتخابی به این سازمان تعلق ندارد.'
        p['eng1'].refresh_from_db()
        assert p['eng1'].advisor_id == before

    def test_moved_student_leaves_the_old_advisors_counts(self, panel):
        """C1 live-scoping: after the move the old advisor's row shrinks."""
        p = panel
        client = _auth(p['manager'])
        before = client.get(f'{ORG_REPORT}?from={FROM}&to={TO}').json()
        assert len(before['advisors']) == 2

        res = client.post(
            REASSIGN.format(pk=p['eng1'].pk),
            {'advisorId': p['adv2'].pk}, format='json',
        )
        assert res.status_code == 200

        after = client.get(f'{ORG_REPORT}?from={FROM}&to={TO}').json()
        rows = {r['advisorName']: r for r in after['advisors']}
        assert rows['آرش کمالی']['studentCount'] == 1   # eng2 only
        assert rows['مریم رضایی']['studentCount'] == 2

    def test_deputy_may_reassign(self, panel):
        p = panel
        deputy = _member(p['org'], OrgRole.DEPUTY, 'deputy')
        res = _auth(deputy).post(
            REASSIGN.format(pk=p['eng3'].pk),
            {'advisorId': p['adv1'].pk}, format='json',
        )
        assert res.status_code == 200

    def test_body_missing_advisorId_message(self, panel):
        res = _auth(panel['manager']).post(
            REASSIGN.format(pk=panel['eng1'].pk), {}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == MSG_BODY

    def test_unknown_advisor_id_message(self, panel):
        res = _auth(panel['manager']).post(
            REASSIGN.format(pk=panel['eng1'].pk),
            {'advisorId': 424242}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'مشاور انتخابی به این سازمان تعلق ندارد.'

    def test_non_advisor_user_rejected(self, panel):
        outsider_student = _student('out-stu', '09120000777')
        res = _auth(panel['manager']).post(
            REASSIGN.format(pk=panel['eng1'].pk),
            {'advisorId': outsider_student.pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'کاربر انتخابی مشاور نیست.'

    def test_advisor_without_membership_message(self, panel):
        """ADVISOR-role user, but never joined this org ⇒ the pinned message."""
        p = panel
        rogue = baker.make(User, username='rogue', role=User.Role.ADVISOR)
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=p['eng1'].pk),
            {'advisorId': rogue.pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'مشاور انتخابی به این سازمان تعلق ندارد.'

    def test_other_orgs_advisor_rejected(self, panel):
        """Member of org B — a different tenancy — is equally foreign here."""
        p = panel
        other = _org(name='دبیرستان رازی')
        foreign_adv = _member(other, OrgRole.ADVISOR)
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=p['eng1'].pk),
            {'advisorId': foreign_adv.pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'مشاور انتخابی به این سازمان تعلق ندارد.'

    def test_same_advisor_message(self, panel):
        res = _auth(panel['manager']).post(
            REASSIGN.format(pk=panel['eng1'].pk),
            {'advisorId': panel['adv1'].pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'این دانش‌آموز از قبل با همین مشاور همکاری می‌کند.'

    def test_ended_engagement_not_movable(self, panel):
        p = panel
        AdvisoryEngagement.objects.filter(pk=p['eng1'].pk).update(status=Status.ENDED)
        res = _auth(p['manager']).post(
            REASSIGN.format(pk=p['eng1'].pk),
            {'advisorId': p['adv2'].pk}, format='json',
        )
        assert res.status_code == 400
        assert res.json()['detail'] == 'فقط همکاری فعال قابل جابجایی است.'

    def test_foreign_engagement_pk_404(self, panel):
        """Another org's engagement id under THIS manager's URL — 404، نه ۴۰۳."""
        p = panel
        other = _org(name='دبیرستان رازی')
        o_adv = _member(other, OrgRole.ADVISOR)
        o_stu = _member(other, OrgRole.STUDENT, phone='09120000888')
        o_eng = _org_engagement(o_adv, o_stu, other)

        res = _auth(p['manager']).post(
            REASSIGN.format(pk=o_eng.pk),
            {'advisorId': p['adv2'].pk}, format='json',
        )
        assert res.status_code == 404
        assert res.json()['detail'] == 'همکاری پیدا نشد.'


# ── risman step 4 head-start: the impersonation claim is a hard gate ──────────

def _impersonated_client(user) -> APIClient:
    """A real JWT whose payload carries the ``imp`` claim (step 4's token).

    Minted exactly the way ``ImpersonationView`` will mint it in step 4 —
    ``RefreshToken.for_user`` plus the injected claim — so these tests pin the
    permission class's behaviour against the *real* token shape, not a mock.
    """
    from rest_framework_simplejwt.tokens import RefreshToken

    refresh = RefreshToken.for_user(user)
    access = refresh.access_token
    access['imp'] = {'by': 999}
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f'Bearer {access}')
    return client


class TestImpersonationTokenHardGate:
    """A member acting under an impersonated token must stay a pure reader.

    These two are the security doctrine's locked doors (گام ۴): no re-chaining
    of impersonations, no reaching the manager panel through the borrowed
    identity. They run against a REAL SimpleJWT token so a future refactor of
    the claim's location fails here loudly instead of silently opening.
    """

    def test_imp_token_cannot_read_the_org_panel(self, panel):
        res = _impersonated_client(panel['adv1']).get(ORG_OVERVIEW)
        assert res.status_code == 403
        assert res.json()['detail'] == 'این جلسه در حالت ورود مستقیم است و این کار مجاز نیست.'

    def test_imp_token_cannot_reassign(self, panel):
        p = panel
        res = _impersonated_client(p['adv1']).post(
            REASSIGN.format(pk=p['eng2'].pk),
            {'advisorId': p['adv2'].pk}, format='json',
        )
        assert res.status_code == 403
        assert res.json()['detail'] == 'این جلسه در حالت ورود مستقیم است و این کار مجاز نیست.'