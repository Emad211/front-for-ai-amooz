"""Risman step 2 — the reporting engine, Excel export, and their endpoints.

Zero-token throughout: every number below is hand-built arithmetic over ORM
fixtures (plans + logs + exam scores), never an LLM call. The doctrine (roadmap
ق۴) is followed exactly:

* exact aggregation math for all three reports, including both clipping rules
  (a plan item dated beyond ``today`` is excluded from planned; a log outside
  the requested range is excluded from actual);
* an empty range answers valid zeros/nulls, never an error;
* reversed / over-long / malformed ranges answer 400 with the pinned Persian
  messages;
* the xlsx branch is **read back** with ``openpyxl.load_workbook`` and its
  sheet, RTL flag and key cells are asserted;
* the full access matrix: owner 200, stranger advisor 404, student 403,
  anonymous 401.

The routes are not wired into ``urls.py`` yet (the orchestrator registers them
after landing), so the view tests mount ``views_reports.urlpatterns`` through
a tiny in-memory urlconf + ``pytest.mark.urls`` — no production file touched.
"""

from __future__ import annotations

import datetime
import io
import sys
import types

import pytest
from django.contrib.auth import get_user_model
from django.urls import include, path
from django.utils import timezone
from model_bakery import baker
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.advisory.models import (
    AdvisoryEngagement,
    DailyLog,
    DailyLogItem,
    StudyExamAnalysis,
    StudyExamScore,
    StudyPlan,
    StudyPlanItem,
    StudentSubject,
    Subject,
    WeeklyAssessment,
)
from apps.advisory.services import excel_export, reports
from apps.advisory import views_reports as reports_views

User = get_user_model()
Status = AdvisoryEngagement.Status
Mode = AdvisoryEngagement.Mode

pytestmark = [pytest.mark.django_db, pytest.mark.api]

PLANNER_URL = '/api/advisory/students/{pk}/reports/planner/'
STUDENT_REPORT_URL = '/api/advisory/students/{pk}/reports/student/'

MSG_BAD_DATE = 'تاریخ باید به شکل YYYY-MM-DD باشد.'
MSG_REVERSED_RANGE = 'تاریخ پایان نمی‌تواند پیش از تاریخ شروع باشد.'
MSG_RANGE_TOO_LONG = 'بازه حداکثر ۹۲ روز است.'
XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

TODAY = timezone.localdate()
FROM = TODAY - datetime.timedelta(days=6)
TO = TODAY


# ── in-memory urlconf so the unwired routes are testable without urls.py ─────

_test_urls = types.ModuleType('_advisory_reports_test_urls')
_test_urls.urlpatterns = [path('api/advisory/', include(reports_views.urlpatterns))]
sys.modules.setdefault('_advisory_reports_test_urls', _test_urls)


def iso(day: datetime.date) -> str:
    return day.isoformat()


# ── fixture helpers ───────────────────────────────────────────────────────────

def _auth(user) -> APIClient:
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def _advisor(username='adv', **kwargs):
    return baker.make(User, username=username, role=User.Role.ADVISOR, **kwargs)


def _student(username='stu', phone='09120000001', **kwargs):
    return baker.make(User, username=username, role=User.Role.STUDENT, phone=phone, **kwargs)


def _engagement(advisor, student, *, status=Status.ACTIVE, **kwargs):
    defaults = {
        'invited_phone': student.phone or '',
        'mode': Mode.FREELANCE,
        'organization': None,
        'status': status,
        'started_on': TODAY - datetime.timedelta(days=60),
    }
    defaults.update(kwargs)
    return AdvisoryEngagement.objects.create(advisor=advisor, student=student, **defaults)


def _subject(name: str) -> Subject:
    return baker.make(Subject, name=name)


def _selection(engagement, subject) -> StudentSubject:
    return baker.make(StudentSubject, engagement=engagement, subject=subject)


def _make_plan(engagement, start, duration_days, rows, status_value) -> StudyPlan:
    """rows = [(day_offset, selection, planned_minutes)] — ORM-direct, no door."""
    plan = baker.make(
        StudyPlan,
        engagement=engagement,
        start_date=start,
        duration_days=duration_days,
        status=status_value,
    )
    StudyPlanItem.objects.bulk_create([
        StudyPlanItem(
            plan=plan,
            day_offset=offset,
            student_subject=selection,
            planned_minutes=minutes,
        )
        for offset, selection, minutes in rows
    ])
    return plan


def _published_plan(engagement, start, duration_days, rows) -> StudyPlan:
    return _make_plan(engagement, start, duration_days, rows, StudyPlan.Status.PUBLISHED)


def _draft_plan(engagement, start, duration_days, rows) -> StudyPlan:
    return _make_plan(engagement, start, duration_days, rows, StudyPlan.Status.DRAFT)


def _log(engagement, log_date, *, tests_taken=0, items=()) -> DailyLog:
    """items = [(selection, actual_minutes)] — one row per (day, subject)."""
    log = baker.make(
        DailyLog,
        engagement=engagement,
        log_date=log_date,
        tests_taken=tests_taken,
    )
    DailyLogItem.objects.bulk_create([
        DailyLogItem(log=log, student_subject=selection, actual_minutes=minutes)
        for selection, minutes in items
    ])
    return log


def _exam_score(engagement, exam_date, title='آزمون جامع', percent='80.00'):
    return baker.make(
        StudyExamScore,
        engagement=engagement,
        title=title,
        exam_kind='SCHOOL',
        exam_date=exam_date,
        score_percent=percent,
    )


def _range_params(**overrides):
    params = {'from': iso(FROM), 'to': iso(TO)}
    params.update(overrides)
    return params


# ── shared fixture graph ──────────────────────────────────────────────────────

@pytest.fixture
def math_world():
    """One engagement with two subjects, a PUBLISHED plan, and three logs.

    Layout (relative to «today» so the clip tests stay deterministic):

    * PUBLISHED plan starting ``FROM``, 9 days long:
      offset 0 → ریاضی ۶۰ (in range), offset 2 → فیزیک ۳۰ (in range),
      offset 8 → ریاضی ۴۵ (**beyond today** ⇒ excluded from planned);
    * a DRAFT plan whose items must never count;
    * logs: ``FROM`` ریاضی ۴۰ + فیزیک ۲۰ (tests ۱۲), ``FROM+1`` ریاضی ۱۰
      (tests ۰), and ``FROM-1`` ریاضی ۹۹ (**outside the range** ⇒ excluded).
    """
    advisor = _advisor()
    student = _student(first_name='زهرا', last_name='محمدی')
    engagement = _engagement(advisor, student)

    math_sel = _selection(engagement, _subject('ریاضی'))
    physics_sel = _selection(engagement, _subject('فیزیک'))

    _published_plan(engagement, FROM, 9, [
        (0, math_sel, 60),
        (2, physics_sel, 30),
        (8, math_sel, 45),  # FROM+8 = today+2 → clipped out of planned
    ])
    _draft_plan(engagement, FROM, 7, [(0, math_sel, 500)])

    _log(engagement, FROM, tests_taken=12, items=[
        (math_sel, 40),
        (physics_sel, 20),
    ])
    _log(engagement, FROM + datetime.timedelta(days=1), tests_taken=0, items=[
        (math_sel, 10),
    ])
    _log(engagement, FROM - datetime.timedelta(days=1), items=[
        (math_sel, 99),
    ])

    return {
        'advisor': advisor,
        'student': student,
        'engagement': engagement,
        'math': math_sel.subject,
        'physics': physics_sel.subject,
    }


# ── planner_report · exact math ───────────────────────────────────────────────

class TestPlannerReportMath:
    def test_days_subjects_and_totals_are_exact(self, math_world):
        report = reports.planner_report(math_world['engagement'], FROM, TO)

        # Days cover the measurable window [FROM, today] ascending.
        assert [day['date'] for day in report['days']] == [
            iso(FROM + datetime.timedelta(days=offset)) for offset in range(7)
        ]
        assert report['days'][0] == {
            'date': iso(FROM), 'planned': 60, 'actual': 60,
        }
        assert report['days'][1] == {
            'date': iso(FROM + datetime.timedelta(days=1)),
            'planned': 0, 'actual': 10,
        }
        assert report['days'][2]['planned'] == 30
        assert report['days'][2]['actual'] == 0
        # No day beyond today is emitted, and nothing else carries data.
        assert len(report['days']) == 7
        assert sum(d['planned'] + d['actual'] for d in report['days'][3:]) == 0

        # Subjects: name-sorted union; future plan minutes excluded.
        by_name = {row['name']: row for row in report['subjects']}
        assert set(by_name) == {'ریاضی', 'فیزیک'}
        assert by_name['ریاضی'] == {
            'subjectId': math_world['math'].pk,
            'name': 'ریاضی',
            'planned': 60,   # the +45 beyond-today row is clipped away
            'actual': 50,    # 40 + 10 inside the range; the -1d 99 is outside
            'coveragePercent': 83,  # round(50/60*100)
        }
        assert by_name['فیزیک']['planned'] == 30
        assert by_name['فیزیک']['actual'] == 20
        assert by_name['فیزیک']['coveragePercent'] == 67  # round(20/30*100)

        assert report['totals'] == {
            'planned': 90, 'actual': 70, 'coveragePercent': 78,
        }

    def test_future_plan_item_is_excluded_from_planned(self, math_world):
        report = reports.planner_report(math_world['engagement'], FROM, TO)
        # Total planned is 60+30 only — the offset-8 row (today+2) never counts.
        assert report['totals']['planned'] == 90

    def test_log_outside_range_is_excluded(self, math_world):
        report = reports.planner_report(math_world['engagement'], FROM, TO)
        # The FROM-1 log (99 minutes) belongs to no measured day here.
        assert report['totals']['actual'] == 70

    def test_empty_range_answers_zeros_and_null_not_error(self):
        advisor = _advisor(username='adv-empty')
        student = _student(username='stu-empty')
        engagement = _engagement(advisor, student)

        report = reports.planner_report(
            engagement, TODAY - datetime.timedelta(days=6), TODAY,
        )

        assert len(report['days']) == 7
        assert all(
            day['planned'] == 0 and day['actual'] == 0
            for day in report['days']
        )
        assert report['subjects'] == []
        assert report['totals'] == {
            'planned': 0, 'actual': 0, 'coveragePercent': None,
        }

    def test_wholly_future_range_yields_no_days(self):
        advisor = _advisor(username='adv-fut')
        student = _student(username='stu-fut')
        engagement = _engagement(advisor, student)

        report = reports.planner_report(
            engagement,
            TODAY + datetime.timedelta(days=1),
            TODAY + datetime.timedelta(days=7),
        )

        assert report['days'] == []
        assert report['totals']['coveragePercent'] is None


# ── student_report · exact math ───────────────────────────────────────────────

class TestStudentReportMath:
    def test_series_share_and_scores_are_exact(self, math_world):
        engagement = math_world['engagement']

        in_range_old = _exam_score(
            engagement, FROM + datetime.timedelta(days=1), 'آزمون قدیمی',
        )
        in_range_new = _exam_score(engagement, TO, 'آزمون جدید')
        _exam_score(engagement, FROM - datetime.timedelta(days=1), 'آزمون بیرون بازه')

        report = reports.student_report(engagement, FROM, TO)

        assert report['studySeries'] == [
            {'date': iso(FROM), 'minutes': 60},
            {'date': iso(FROM + datetime.timedelta(days=1)), 'minutes': 10},
        ]
        # Only days with tests_taken > 0 appear.
        assert report['testSeries'] == [{'date': iso(FROM), 'testsTaken': 12}]

        # Descending by minutes; shares rounded to 1 decimal of the grand 70.
        assert [(row['name'], row['minutes'], row['sharePercent'])
                for row in report['subjectShare']] == [
            ('ریاضی', 50, 71.4),
            ('فیزیک', 20, 28.6),
        ]

        # Newest exam first; the out-of-range row is gone.
        assert [row.pk for row in report['examScores']] == [
            in_range_new.pk, in_range_old.pk,
        ]

    def test_empty_range_answers_empty_lists(self):
        advisor = _advisor(username='adv-s')
        student = _student(username='stu-s')
        engagement = _engagement(advisor, student)

        report = reports.student_report(
            engagement, TODAY - datetime.timedelta(days=6), TODAY,
        )

        assert report == {
            'studySeries': [],
            'testSeries': [],
            'subjectShare': [],
            'examScores': [],
        }


# ── advisor_report · exact math ───────────────────────────────────────────────

class TestAdvisorReportMath:
    def test_per_student_rows_and_tool_counters(self, math_world):
        advisor = math_world['advisor']
        other_student = _student(username='stu-b', phone='09120000002')
        other_engagement = _engagement(advisor, other_student)  # no data at all

        # Tool counters: one assessment anchored on the window's Saturday, one
        # analysis created now, plus negatives (draft plan, out-of-range rows).
        saturday = FROM + datetime.timedelta(days=(5 - FROM.weekday()) % 7)
        baker.make(
            WeeklyAssessment,
            engagement=math_world['engagement'],
            week_start=saturday,
            scores={'focus': 4},
        )
        baker.make(
            StudyExamAnalysis,
            engagement=math_world['engagement'],
            exam_date=FROM,
        )
        stale_analysis = baker.make(
            StudyExamAnalysis,
            engagement=math_world['engagement'],
            exam_date=FROM,
        )
        stale_analysis.created_at = timezone.now() - datetime.timedelta(days=400)
        stale_analysis.save(update_fields=['created_at'])

        report = reports.advisor_report(advisor, FROM, TO)

        rows = {row['engagementId']: row for row in report['students']}
        assert set(rows) == {math_world['engagement'].pk, other_engagement.pk}
        active_row = rows[math_world['engagement'].pk]
        assert active_row['studentName'] == 'زهرا محمدی'
        assert active_row['planned'] == 90
        assert active_row['actual'] == 70
        assert active_row['coveragePercent'] == 78
        assert active_row['testsTaken'] == 12

        empty_row = rows[other_engagement.pk]
        assert empty_row['planned'] == 0
        assert empty_row['actual'] == 0
        assert empty_row['coveragePercent'] is None
        assert empty_row['testsTaken'] == 0

        assert report['tools'] == {
            'plansPublished': 1,      # the PUBLISHED plan; the draft never counts
            'assessmentsWritten': 1,  # the stale analysis is out of range
            'analysesCreated': 1,
        }


# ── excel export · build + read-back ─────────────────────────────────────────

class TestExcelExport:
    def test_planner_workbook_round_trips(self, math_world):
        report = reports.planner_report(math_world['engagement'], FROM, TO)
        buffer = excel_export.report_workbook('planner', report)

        workbook = load_workbook(io.BytesIO(buffer.getvalue()))
        sheet = workbook['برنامه']
        assert sheet.sheet_view.rightToLeft is True

        assert [sheet.cell(row=1, column=c).value for c in range(1, 6)] == [
            'تاریخ', 'روز', 'برنامه‌ریزی‌شده', 'انجام‌شده', 'پوشش٪',
        ]
        # First day row: ISO date, Persian weekday, planned 60, actual 60.
        assert sheet.cell(row=2, column=1).value == iso(FROM)
        assert sheet.cell(row=2, column=2).value != ''
        assert sheet.cell(row=2, column=3).value == 60
        assert sheet.cell(row=2, column=4).value == 60

        # Totals line closes the sheet: جمع | — | 90 | 70 | 78.
        last_row = sheet.max_row
        assert sheet.cell(row=last_row, column=1).value == 'جمع'
        assert sheet.cell(row=last_row, column=3).value == 90
        assert sheet.cell(row=last_row, column=4).value == 70
        assert sheet.cell(row=last_row, column=5).value == 78

    def test_unknown_kind_raises_value_error(self):
        with pytest.raises(ValueError):
            excel_export.report_workbook('mystery', {})


# ── endpoints · JSON, xlsx, validation, access matrix ────────────────────────

@pytest.mark.urls('_advisory_reports_test_urls')
class TestPlannerReportEndpoint:
    def test_owner_gets_exact_wire_payload(self, math_world):
        response = _auth(math_world['advisor']).get(
            PLANNER_URL.format(pk=math_world['engagement'].pk), _range_params(),
        )

        assert response.status_code == 200
        assert response.data['totals'] == {
            'planned': 90, 'actual': 70, 'coveragePercent': 78,
        }
        assert response.data['days'][0] == {
            'date': iso(FROM), 'planned': 60, 'actual': 60,
        }
        assert {row['name'] for row in response.data['subjects']} == {'ریاضی', 'فیزیک'}

    def test_xlsx_branch_serves_a_readable_workbook(self, math_world):
        response = _auth(math_world['advisor']).get(
            PLANNER_URL.format(pk=math_world['engagement'].pk),
            _range_params(format='xlsx'),
        )

        assert response.status_code == 200
        assert response['Content-Type'] == XLSX_CONTENT_TYPE
        assert response['Content-Disposition'] == (
            f'attachment; filename=report-planner-{iso(FROM)}_{iso(TO)}.xlsx'
        )

        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook['برنامه']
        assert sheet.sheet_view.rightToLeft is True
        assert sheet.cell(row=1, column=1).value == 'تاریخ'
        assert sheet.cell(row=2, column=3).value == 60

    def test_reversed_range_is_400(self, math_world):
        response = _auth(math_world['advisor']).get(
            PLANNER_URL.format(pk=math_world['engagement'].pk),
            {'to': iso(FROM), 'from': iso(TO)},
        )
        assert response.status_code == 400
        assert response.data['detail'] == MSG_REVERSED_RANGE

    def test_over_long_range_is_400(self, math_world):
        response = _auth(math_world['advisor']).get(
            PLANNER_URL.format(pk=math_world['engagement'].pk),
            {'from': iso(TODAY - datetime.timedelta(days=92)), 'to': iso(TODAY)},
        )
        assert response.status_code == 400
        assert response.data['detail'] == MSG_RANGE_TOO_LONG

    def test_exactly_ninety_two_days_passes(self, math_world):
        response = _auth(math_world['advisor']).get(
            PLANNER_URL.format(pk=math_world['engagement'].pk),
            {'from': iso(TODAY - datetime.timedelta(days=91)), 'to': iso(TODAY)},
        )
        assert response.status_code == 200

    def test_malformed_date_is_400(self, math_world):
        response = _auth(math_world['advisor']).get(
            PLANNER_URL.format(pk=math_world['engagement'].pk),
            {'from': '2026-8-01', 'to': iso(TODAY)},
        )
        assert response.status_code == 400
        assert response.data['detail'] == MSG_BAD_DATE

    def test_missing_dates_are_400(self, math_world):
        response = _auth(math_world['advisor']).get(
            PLANNER_URL.format(pk=math_world['engagement'].pk),
        )
        assert response.status_code == 400
        assert response.data['detail'] == MSG_BAD_DATE


@pytest.mark.urls('_advisory_reports_test_urls')
class TestStudentReportEndpoint:
    def test_owner_gets_the_full_shape(self, math_world):
        _exam_score(math_world['engagement'], FROM, 'آزمون در بازه')

        response = _auth(math_world['advisor']).get(
            STUDENT_REPORT_URL.format(pk=math_world['engagement'].pk),
            _range_params(),
        )

        assert response.status_code == 200
        assert response.data['studySeries'][0] == {
            'date': iso(FROM), 'minutes': 60,
        }
        assert response.data['testSeries'] == [{'date': iso(FROM), 'testsTaken': 12}]
        assert response.data['subjectShare'][0]['name'] == 'ریاضی'
        assert len(response.data['examScores']) == 1
        assert response.data['examScores'][0]['title'] == 'آزمون در بازه'
        assert response.data['examScores'][0]['scorePercent'] == 80.0

    def test_reversed_range_is_400(self, math_world):
        response = _auth(math_world['advisor']).get(
            STUDENT_REPORT_URL.format(pk=math_world['engagement'].pk),
            {'to': iso(FROM), 'from': iso(TO)},
        )
        assert response.status_code == 400
        assert response.data['detail'] == MSG_REVERSED_RANGE


@pytest.mark.urls('_advisory_reports_test_urls')
class TestReportAccessMatrix:
    """owner 200 · stranger advisor 404 · student 403 · anonymous 401."""

    @pytest.mark.parametrize('url_template', [PLANNER_URL, STUDENT_REPORT_URL])
    def test_matrix(self, math_world, url_template):
        owner = math_world['advisor']
        stranger = _advisor(username='stranger')
        outsider_student = _student(username='outsider', phone='09120000003')

        url = url_template.format(pk=math_world['engagement'].pk)

        assert _auth(owner).get(url, _range_params()).status_code == 200
        assert _auth(stranger).get(url, _range_params()).status_code == 404
        assert _auth(outsider_student).get(url, _range_params()).status_code == 403
        assert APIClient().get(url, _range_params()).status_code == 401
